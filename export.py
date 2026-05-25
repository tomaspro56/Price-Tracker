"""
Módulo de exportación del price tracker.

Responsabilidad única: tomar datos de database.py y escribirlos
a archivos Excel o CSV. No sabe nada de scraping ni de la web.
"""

import csv
import logging
import os
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import database as db

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Constantes de estilo Excel ────────────────────────────────────────────────
COLOR_ENCABEZADO = "1E3A5F"   # azul oscuro (mismo que el dashboard)
COLOR_TEXTO_ENC  = "FFFFFF"   # blanco

FUENTE_ENCABEZADO = Font(bold=True, color=COLOR_TEXTO_ENC, size=11)
RELLENO_ENCABEZADO = PatternFill(fill_type="solid", fgColor=COLOR_ENCABEZADO)
ALINEACION_CENTRO = Alignment(horizontal="center", vertical="center")
ALINEACION_IZQ    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

CARPETA_EXPORTS = "exports"


# ── Utilidades internas ───────────────────────────────────────────────────────

def _asegurar_carpeta_exports() -> None:
    """Crea la carpeta de exports si no existe."""
    os.makedirs(CARPETA_EXPORTS, exist_ok=True)


def _timestamp_nombre() -> str:
    """Sufijo de timestamp para nombres de archivo: '20240115_103045'."""
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _aplicar_estilo_encabezado(hoja, fila_encabezados: list[str]) -> None:
    """Escribe y estiliza la fila de encabezados en la hoja activa."""
    for col, nombre in enumerate(fila_encabezados, start=1):
        celda = hoja.cell(row=1, column=col, value=nombre)
        celda.font = FUENTE_ENCABEZADO
        celda.fill = RELLENO_ENCABEZADO
        celda.alignment = ALINEACION_CENTRO
    hoja.row_dimensions[1].height = 22


def _ajustar_anchos(hoja, anchos_minimos: dict[int, int]) -> None:
    """
    Ajusta el ancho de cada columna al contenido más largo,
    respetando un mínimo configurable por columna.
    """
    for col_idx, min_ancho in anchos_minimos.items():
        letra = get_column_letter(col_idx)
        ancho_maximo = min_ancho
        for celda in hoja[letra]:
            if celda.value:
                largo = len(str(celda.value))
                ancho_maximo = max(ancho_maximo, largo)
        hoja.column_dimensions[letra].width = min(ancho_maximo + 2, 60)


# ── API pública ───────────────────────────────────────────────────────────────

def exportar_catalogo_excel(
    ruta_salida: str = "",
    ruta_db: str = "precios.db",
) -> str:
    """
    Exporta el catálogo de libros (un registro por libro, precio actual)
    a un archivo Excel con encabezados estilizados.

    Args:
        ruta_salida: Ruta del archivo de salida. Si está vacía, genera un
                     nombre automático con timestamp en la carpeta exports/.
        ruta_db: Ruta al archivo SQLite.

    Returns:
        Ruta absoluta del archivo generado.

    Raises:
        ValueError: Si la base de datos no tiene libros.
    """
    _asegurar_carpeta_exports()
    if not ruta_salida:
        ruta_salida = os.path.join(CARPETA_EXPORTS, f"catalogo_{_timestamp_nombre()}.xlsx")

    libros = db.obtener_libros_con_precio_actual(ruta_db)
    if not libros:
        raise ValueError("La base de datos no tiene libros. Ejecutá una actualización primero.")

    libro_excel = openpyxl.Workbook()
    hoja = libro_excel.active
    hoja.title = "Catálogo"

    encabezados = ["Título", "Precio (£)", "Rating", "Disponibilidad", "Última actualización"]
    _aplicar_estilo_encabezado(hoja, encabezados)

    for fila_num, libro in enumerate(libros, start=2):
        hoja.cell(row=fila_num, column=1, value=libro["titulo"]).alignment = ALINEACION_IZQ
        # Precio como número para que Excel pueda operar con él
        celda_precio = hoja.cell(row=fila_num, column=2, value=round(libro["precio_actual"], 2))
        celda_precio.number_format = '£#,##0.00'
        celda_precio.alignment = ALINEACION_CENTRO

        hoja.cell(row=fila_num, column=3, value=libro["rating"]).alignment = ALINEACION_CENTRO
        hoja.cell(row=fila_num, column=4, value=libro["disponibilidad"]).alignment = ALINEACION_CENTRO
        # Limpiar el timestamp ISO para que sea legible
        fecha_limpia = libro["ultima_actualizacion"][:16].replace("T", " ")
        hoja.cell(row=fila_num, column=5, value=fecha_limpia).alignment = ALINEACION_CENTRO

    _ajustar_anchos(hoja, {1: 35, 2: 12, 3: 8, 4: 15, 5: 20})

    # Fijar la fila de encabezados al hacer scroll
    hoja.freeze_panes = "A2"

    libro_excel.save(ruta_salida)
    logger.info("Excel catálogo exportado: %s (%d libros)", ruta_salida, len(libros))
    return os.path.abspath(ruta_salida)


def exportar_catalogo_csv(
    ruta_salida: str = "",
    ruta_db: str = "precios.db",
) -> str:
    """
    Exporta el catálogo de libros a CSV con encoding utf-8-sig para que
    Excel lo abra correctamente con tildes y caracteres especiales.

    Args:
        ruta_salida: Ruta del archivo de salida. Si está vacía, genera un
                     nombre automático con timestamp en la carpeta exports/.
        ruta_db: Ruta al archivo SQLite.

    Returns:
        Ruta absoluta del archivo generado.

    Raises:
        ValueError: Si la base de datos no tiene libros.
    """
    _asegurar_carpeta_exports()
    if not ruta_salida:
        ruta_salida = os.path.join(CARPETA_EXPORTS, f"catalogo_{_timestamp_nombre()}.csv")

    libros = db.obtener_libros_con_precio_actual(ruta_db)
    if not libros:
        raise ValueError("La base de datos no tiene libros. Ejecutá una actualización primero.")

    # utf-8-sig agrega el BOM que necesita Excel para detectar el encoding
    with open(ruta_salida, "w", newline="", encoding="utf-8-sig") as archivo:
        escritor = csv.writer(archivo)
        escritor.writerow(["Título", "Precio (£)", "Rating", "Disponibilidad", "Última actualización"])
        for libro in libros:
            escritor.writerow([
                libro["titulo"],
                f"{libro['precio_actual']:.2f}",
                libro["rating"],
                libro["disponibilidad"],
                libro["ultima_actualizacion"][:16].replace("T", " "),
            ])

    logger.info("CSV catálogo exportado: %s (%d libros)", ruta_salida, len(libros))
    return os.path.abspath(ruta_salida)


def exportar_historial_excel(
    ruta_salida: str = "",
    ruta_db: str = "precios.db",
) -> str:
    """
    Exporta el historial completo de precios (todos los snapshots de todos
    los libros) a Excel. Útil para análisis de tendencias externo.

    Args:
        ruta_salida: Ruta del archivo de salida. Si está vacía, genera un
                     nombre automático con timestamp en la carpeta exports/.
        ruta_db: Ruta al archivo SQLite.

    Returns:
        Ruta absoluta del archivo generado.

    Raises:
        ValueError: Si la base de datos no tiene libros o historial.
    """
    _asegurar_carpeta_exports()
    if not ruta_salida:
        ruta_salida = os.path.join(CARPETA_EXPORTS, f"historial_{_timestamp_nombre()}.xlsx")

    # Necesitamos cruzar libros con su historial; obtenemos el catálogo
    # para tener los títulos y luego pedimos el historial de cada uno.
    libros = db.obtener_libros_con_precio_actual(ruta_db)
    if not libros:
        raise ValueError("La base de datos no tiene libros. Ejecutá una actualización primero.")

    libro_excel = openpyxl.Workbook()
    hoja = libro_excel.active
    hoja.title = "Historial completo"

    encabezados = ["Título", "Precio (£)", "Disponibilidad", "Fecha (UTC)"]
    _aplicar_estilo_encabezado(hoja, encabezados)

    fila_num = 2
    total_snapshots = 0

    for libro in libros:
        historial = db.obtener_historial_libro(libro["id"], ruta_db)
        for snap in historial:
            hoja.cell(row=fila_num, column=1, value=libro["titulo"]).alignment = ALINEACION_IZQ
            celda_precio = hoja.cell(row=fila_num, column=2, value=round(snap["precio"], 2))
            celda_precio.number_format = '£#,##0.00'
            celda_precio.alignment = ALINEACION_CENTRO
            hoja.cell(row=fila_num, column=3, value=snap["disponibilidad"]).alignment = ALINEACION_CENTRO
            fecha_limpia = snap["fecha"][:16].replace("T", " ")
            hoja.cell(row=fila_num, column=4, value=fecha_limpia).alignment = ALINEACION_CENTRO
            fila_num += 1
            total_snapshots += 1

    if total_snapshots == 0:
        raise ValueError("No hay snapshots en el historial todavía.")

    _ajustar_anchos(hoja, {1: 35, 2: 12, 3: 15, 4: 20})
    hoja.freeze_panes = "A2"

    libro_excel.save(ruta_salida)
    logger.info(
        "Excel historial exportado: %s (%d snapshots, %d libros)",
        ruta_salida, total_snapshots, len(libros),
    )
    return os.path.abspath(ruta_salida)
